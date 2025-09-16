from flask import Flask, render_template, request, send_file
import os #os path 경로
from openpyxl import load_workbook #엑셀파일 제어
from deep_translator import GoogleTranslator #번역


app = Flask(__name__)

@app.route("/") 
def index(): 
    return render_template('index.html')
    
@app.route("/upload", methods=['GET', 'POST']) 
def upload(): 
    file = request.files["file"] 
    #"uploads" 라는 폴더 안에, 업로드된 파일의 원래 이름(file.filename)을 붙여서 경로를 만듦
    #uploads/파일명
    file.save(os.path.join("uploads", file.filename))

    workbook = load_workbook(os.path.join("uploads", file.filename))
    sheet = workbook.active #엑셀 파일안에서 현재 활성화된 시트를 불러옴
    for row in sheet.iter_rows(): ###row(행)의 데이터가 없을때까지 range대신에 사용
        for cell in row: #셀 value값을 row에서 끄집어내기
            translated_text = GoogleTranslator(source='ko', target='en').translate(cell.value) #셀 value값을 번역
            cell.value = translated_text #번역 결과를 셀값으로 저장
    workbook.save('result_en.xlsx')         

    return render_template('result.html')

#번역결과를 웹사이트에서 보여주기--> 별도 공부!
#번역결과에 개인정보가 있는지 탐지하는 것도 추가! --> 별도공부!!
#번역결과를 다운로드 할 수 있게 도와주기

@app.route("/download_report")
def download_report():
    #from flask에 import send_file 모듈 추가
    #사용자에게 result_en.xlsx파일을 보낸다.
    return send_file('result_en.xlsx', as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True) 
    