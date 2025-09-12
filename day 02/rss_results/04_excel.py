from openpyxl import load_workbook

wb = load_workbook("excel_data2.xlsx", data_only = True) #data값만 가져옴
ws = wb.active

cell = ws["A1:E7"]
for row in cell:
    result = [] #저장을 해서 프린트할것, 보기좋게 하기위해
    for cell in row:
        result.append(cell.value)
    print(result)


