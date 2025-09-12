import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook


url = "https://www.malware-traffic-analysis.net/2023/index.html"

header_info = {'User-agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36'}

r = requests.get(url, headers=header_info) 

soup = BeautifulSoup(r.text, 'html.parser')
tags = soup.select("#main_content > div.content > ul > li > a.main_menu")

#퀴즈
#malware-traffic-analysis 수집하여 아래코드를 이용해 엑셀로 저장
#txt파일 저장을 엑셀 파일 형태로 저장하는 방식

#워크북
wb = Workbook() 
ws = wb.active 

ws['A1'] = "설명"
ws['B1'] = "URL 링크"

#malware-traffic-analysis 수집하여 아래코드를 이용해 엑셀로 저장
#txt파일 저장을 엑셀 파일 형태로 저장하는 방식

i = 2
for tag in tags:
    ws.cell(row=i, column=1, value=tag.text)
    ws.cell(row=i, column=2, value=f"https://www.malware-traffic-analysis.net/2023/{tag.get('href')}")
    i = i + 1


#i값없이 appen을 이용해서 하나씩 쌓는 방법으로 이렇게도 구현 가능
#for tag in tags:
#    ws.append([tag.text, f"https://www.malware-traffic-analysis.net/2023/{tag.get('href')}"])




#엑셀 저장(덮어쓰기)
wb.save('Malware.xlsx')