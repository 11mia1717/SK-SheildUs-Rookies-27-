from openpyxl import Workbook

wb = Workbook() #워크북생성
ws = wb.active #워크시트에 접근
ws.title = "새 워크시트 타이틀"

"""
#엑셀 셀에 데이터 넣기
#셀 범위접근
for row in ws['A1:D2']: #행 쉘 순서로 작성, a1부터 d4까지
    for cell in row:
        cell.value = "굿모닝!!" #셀에 데이터 넣기
"""

for i in range(10):
    row_cell = ws.cell(row=(i+1), column=1)
    row_cell.value = str(i+1) + " 번째 데이터 저장"


#엑셀 저장(덮어쓰기)
wb.save('text.xlsx')