"""
퀴즈
1. RSS 서비스에서 list.txt 파일의 정보를 수집한다.  (주의 : 데이터가 매칭이 안되면 제외)
2. Feedparser를 이용해서 정보를 수집
2. 수집된 정보를 xlsx 파일에 리스트별로 저장(3개 파일)한다.
3. results 폴더에 3개의 xlsx 파일이 저장이 될 것이다.
4. 결과 값 디렉터리를 압축을 한다.
5. zip 파일을 FTP에서 전송한다.

list.txt 파일 내용
https://www.dailysecu.com/rss/allArticle.xml
https://www.dailysecu.com/rss/S1N2.xml
https://www.dailysecu.com/rss/clickTop.xml
"""


"""
import feedparser
from openpyxl import Workbook


#url을 list파일에서 받아오기
urls = []

with open('list.txt', 'r', encoding='utf-8') as readFile:
    urls = readFile.readlines()   


#Feedparser를 이용해서 정보를 수집
wb = Workbook()
ws = wb.active

for url in urls:

    feed = feedparser.parse(url)

    titles = []
    links = []
    descriptions = []
    authors = []
    pubDates = []

    for entry in feed.entries:
        titles.append(entry.title)
        links.append(entry.link)
        descriptions.append(entry.description)
        authors.append(entry.author)
        pubDates.append(entry.published)

    data = {"제목": titles, "링크": links, "요약": descriptions, "작성자": authors, "날짜": pubDates}

    print("데이터 생성")

#수집된 정보를 xlsx 파일에 리스트별로 저장(3개 파일)한다.
wb = Workbook()
ws = wb.active

ws['A1'] = "제목"
ws['B1'] = "링크"
ws['C1'] = "요약"
ws['D1'] = "작성자"
ws['E1'] = "날짜"

for url in tags:
    ws.append([tag.text, f"https://www.malware-traffic-analysis.net/2023/{tag.get('href')}" ])
               
               
"""     
               
import feedparser
from openpyxl import Workbook
import os
import zipfile
import ftplib

##RSS 서비스 대상으로 정보 수집 및 엑셀 파일 저장
RESULT_DIR = "results"

#RESULT 폴더 생성
if not os.path.exists(RESULT_DIR):
    os.makedirs(RESULT_DIR)

#list 파일에서 정보 가져오기
with open('list.txt', 'r', encoding='UTF-8') as file:
    rss_urls = file.readlines() #라인을 한줄씩 가져옴

    for index, url in enumerate(rss_urls): #리스트 순회할 때, 인덱스 번호와 값을 동시에 
        feed = feedparser.parse(url)
        titles = []
        links = []
        descriptions = []
        authors = []
        pubDates = []

        for entry in feed.entries:
            titles.append(entry.title)
            links.append(entry.link)
            descriptions.append(entry.description)
            authors.append(entry.author)
            pubDates.append(entry.published)

        # 워크북과 워크시트를 생성합니다.
        wb = Workbook()
        ws = wb.active
        ws.title = f"{index+1}번째 Data" #워크시트 타이틀

        # 첫 번째 행에 헤더를 작성합니다.
        headers = ['제목', '링크', '요약', '작성자', '날짜']
        ws.append(headers)

        # 데이터를 행 단위로 작성합니다.
        for i in range(len(titles)):
            ws.append([titles[i], links[i], descriptions[i], authors[i], pubDates[i]])

        # 엑셀 파일로 저장합니다.
        file_path = os.path.join(RESULT_DIR, f'{index+1}_result.xlsx')
        ##여러 개 문자열을 현재 운영체제에 맞는 경로 문자열로 합쳐줌.
        #단순히 "+" 로 문자열을 더하는 것보다 안전
        wb.save(file_path)


##결과 값을 result.zip 파일로 저장
zip_file = zipfile.ZipFile("result.zip", "w")

for root, dirs, files in os.walk(RESULT_DIR):
    for file in files:
        zip_file.write(os.path.join(root, file))
        #여러 개 문자열을 현재 운영체제에 맞는 경로 문자열로 합쳐줌.
        #단순히 "+" 로 문자열을 더하는 것보다 안전

zip_file.close()

##FTP 서버에 결과 값을 보낸다.
hostname = "192.168.18.128"
ftp = ftplib.FTP(hostname)
ftp.login('msfadmin','msfadmin')
ftp.retrlines('LIST')

with open("result.zip", "rb") as f:
    ftp.storbinary(f"STOR result.zip",f)

print(f"현재작업디렉터리：{ftp.pwd()}")
ftp.retrlines('LIST')
ftp.quit()
